from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class ProjectTemplatePaths:
    output: Path


PROJECT_TYPES = ["delivery", "rnd"]
PROJECT_ROLES = ["PO", "TO", "SM", "TL", "PM", "member"]


def export_project_collection_xlsx(output: Path, *, db_path: Path | None = None) -> tuple[Path, dict[str, Any]]:
    """
    Create an Excel workbook for collecting project master data and mappings.
    The workbook is designed to be edited manually, then saved/exported to CSVs.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.datavalidation import DataValidation

    out = output.expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    note_align = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.alignment = header_align
        ws.freeze_panes = "A2"

    # Sheet 0: README
    ws0 = wb.active
    ws0.title = "README"
    ws0["A1"] = "AskTony 项目维度信息采集模板"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A3"] = "说明："
    ws0["A4"] = "1) 在各 sheet 中填写数据。"
    ws0["A5"] = "2) 直接导入这个 XLSX（推荐，先 dry-run）："
    ws0["A6"] = "   asktony import-project-info --input output/project_info_collection.xlsx --dry-run"
    ws0["A7"] = "3) 通过后写入："
    ws0["A8"] = "   asktony import-project-info --input output/project_info_collection.xlsx"
    ws0["A9"] = "4) 然后运行 model build，再做项目分析："
    ws0["A10"] = "   asktony model build"
    ws0["A11"] = "   asktony analyze project-activity --months 2 --all --csv output/project_activity.csv"
    ws0["A13"] = "备注：如需从表格系统/协作工具流转，也可以把各 sheet 单独导出为 CSV，但不是必须。"
    ws0["A15"] = "字段约定："
    ws0["A16"] = "- project_id：可选；为空时导入会尝试用系统转写从 project_name 生成拼音 id"
    ws0["A17"] = "- 日期字段：YYYY-MM-DD；end_at 可为空表示一直有效"
    ws0["A18"] = "- weight/allocation：0~1；不填默认 1"
    ws0["A19"] = "- 映射 sheet 中 project_name：推荐填写（下拉选择），导入时会自动生成对应的 project_id"
    ws0["A20"] = "- 成员 sheet 推荐按 full_name 选择，employee_id 会自动匹配（也可手填 employee_id）"
    ws0.column_dimensions["A"].width = 90
    for r in range(3, 30):
        ws0.cell(row=r, column=1).alignment = note_align

    # Optional lookup data from DB
    lookup_info: dict[str, Any] = {"repos_rows": 0, "employees_rows": 0, "lookup_error": ""}
    repos_lookup: list[tuple[Any, ...]] = []
    employees_lookup: list[tuple[Any, ...]] = []
    if db_path is not None:
        try:
            from asktony.db import DB

            with DB(db_path).connect() as conn:
                # repos: prefer enriched dim_repo, fallback to base/silver
                for sql in [
                    "SELECT repo_id, repo_name, repo_path FROM gold.dim_repo ORDER BY repo_id",
                    "SELECT repo_id, repo_name, repo_path FROM gold.dim_repo_base ORDER BY repo_id",
                    "SELECT repo_id, repo_name, repo_path FROM silver.repos ORDER BY repo_id",
                ]:
                    try:
                        repos_lookup = conn.execute(sql).fetchall()
                        break
                    except Exception:  # noqa: BLE001
                        continue

                # employees: prefer enrichment table (doesn't require dim_member view)
                for sql in [
                    """
                    SELECT DISTINCT
                      e.employee_id,
                      COALESCE(e.full_name, '') AS full_name,
                      COALESCE(e.line_manager, '') AS line_manager,
                      COALESCE(e.role, '') AS role,
                      COALESCE(d2.name, '') AS department_level2_name,
                      COALESCE(d3.name, '') AS department_level3_name
                    FROM gold.dim_member_enrichment e
                    LEFT JOIN gold.dim_department_level3 d3 ON d3.department_level3_id = e.department_level3_id
                    LEFT JOIN gold.dim_department_level2 d2
                      ON d2.department_level2_id = COALESCE(e.department_level2_id, d3.department_level2_id)
                    WHERE e.employee_id IS NOT NULL AND e.employee_id <> ''
                    ORDER BY e.employee_id
                    """,
                    """
                    SELECT DISTINCT
                      employee_id,
                      COALESCE(full_name, '') AS full_name,
                      COALESCE(line_manager, '') AS line_manager,
                      COALESCE(role, '') AS role,
                      COALESCE(department_level2_name, '') AS department_level2_name,
                      COALESCE(department_level3_name, '') AS department_level3_name
                    FROM gold.dim_member
                    WHERE employee_id IS NOT NULL AND employee_id <> ''
                    ORDER BY employee_id
                    """,
                ]:
                    try:
                        employees_lookup = conn.execute(sql).fetchall()
                        break
                    except Exception:  # noqa: BLE001
                        continue
        except Exception as e:  # noqa: BLE001
            lookup_info["lookup_error"] = str(e)

    lookup_info["repos_rows"] = int(len(repos_lookup))
    lookup_info["employees_rows"] = int(len(employees_lookup))
    if lookup_info["lookup_error"]:
        ws0["A21"] = f"lookup 加载失败：{lookup_info['lookup_error']}"
    else:
        ws0["A21"] = f"lookup 加载结果：repos={lookup_info['repos_rows']} 行，employees={lookup_info['employees_rows']} 行"
    ws0["A21"].alignment = note_align

    # Sheet: repos_lookup
    ws_repo = wb.create_sheet("repos_lookup")
    ws_repo.append(["repo_id", "repo_name", "repo_path"])
    style_header(ws_repo, 3)
    ws_repo.column_dimensions["A"].width = 44
    ws_repo.column_dimensions["B"].width = 26
    ws_repo.column_dimensions["C"].width = 44
    for r in repos_lookup[:5000]:
        ws_repo.append(list(r))

    # Sheet: employees_lookup
    ws_emp = wb.create_sheet("employees_lookup")
    ws_emp.append(
        ["employee_id", "full_name", "line_manager", "role", "department_level2_name", "department_level3_name"]
    )
    style_header(ws_emp, 6)
    ws_emp.column_dimensions["A"].width = 14
    ws_emp.column_dimensions["B"].width = 14
    ws_emp.column_dimensions["C"].width = 14
    ws_emp.column_dimensions["D"].width = 14
    ws_emp.column_dimensions["E"].width = 18
    ws_emp.column_dimensions["F"].width = 18
    for r in employees_lookup[:10000]:
        ws_emp.append(list(r))

    # Sheet 1: dim_project
    ws1 = wb.create_sheet("dim_project")
    cols1 = ["project_name", "project_id", "project_type", "status"]
    ws1.append(cols1)
    style_header(ws1, len(cols1))
    ws1.auto_filter.ref = "A1:D1"
    ws1.append(["农业交付", "", "delivery", "active"])
    ws1.append(["行业大模型", "", "rnd", "active"])
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 14

    dv_type = DataValidation(type="list", formula1=f'"{",".join(PROJECT_TYPES)}"', allow_blank=True)
    ws1.add_data_validation(dv_type)
    dv_type.add("C2:C10000")

    # Sheet 2: bridge_project_repo
    ws2 = wb.create_sheet("bridge_project_repo")
    cols2 = ["project_name", "repo_id", "start_at", "end_at", "weight"]
    ws2.append(cols2)
    style_header(ws2, len(cols2))
    ws2.auto_filter.ref = "A1:E1"
    ws2.append(["农业交付", "clife/farm/server/server-farm-platform", "2026-01-01", "", 1])
    ws2.append(["行业大模型", "clife/ai/modeling/repo", "2026-01-01", "", 0.5])
    ws2.append(["行业大模型", "clife/shared/platform/repo", "2026-01-01", "", 0.5])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 10

    dv_proj_name = DataValidation(type="list", formula1="=dim_project!$A$2:$A$10000", allow_blank=False)
    ws2.add_data_validation(dv_proj_name)
    dv_proj_name.add("A2:A10000")

    dv_repo = DataValidation(type="list", formula1="=repos_lookup!$A$2:$A$5001", allow_blank=False)
    ws2.add_data_validation(dv_repo)
    dv_repo.add("B2:B10000")

    dv_weight = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    ws2.add_data_validation(dv_weight)
    dv_weight.add("E2:E10000")

    # Sheet 3: bridge_project_person_role
    ws3 = wb.create_sheet("bridge_project_person_role")
    cols3 = ["project_name", "full_name", "employee_id", "project_role", "start_at", "end_at", "allocation"]
    ws3.append(cols3)
    style_header(ws3, len(cols3))
    ws3.auto_filter.ref = "A1:G1"
    # 示例：employee_id 列用公式从 full_name 匹配（也允许手填覆盖）
    ws3.append(["农业交付", "张三", "", "TL", "2026-01-01", "", 1])
    ws3.append(["农业交付", "徐鹏", "", "member", "2026-01-01", "", 0.5])
    ws3.append(["行业大模型", "徐鹏", "", "member", "2026-01-01", "", 0.5])
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 10

    dv_proj_name2 = DataValidation(type="list", formula1="=dim_project!$A$2:$A$10000", allow_blank=False)
    ws3.add_data_validation(dv_proj_name2)
    dv_proj_name2.add("A2:A10000")

    dv_full_name = DataValidation(type="list", formula1="=employees_lookup!$B$2:$B$10001", allow_blank=False)
    ws3.add_data_validation(dv_full_name)
    dv_full_name.add("B2:B10000")

    dv_role = DataValidation(type="list", formula1=f'"{",".join(PROJECT_ROLES)}"', allow_blank=True)
    ws3.add_data_validation(dv_role)
    dv_role.add("D2:D10000")

    dv_alloc = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    ws3.add_data_validation(dv_alloc)
    dv_alloc.add("G2:G10000")

    # Auto-fill employee_id from full_name (MATCH/INDEX for broad compatibility)
    for r in range(2, 5000):
        ws3.cell(row=r, column=3).value = (
            f'=IFERROR(INDEX(employees_lookup!$A:$A, MATCH($B{r}, employees_lookup!$B:$B, 0)), "")'
        )

    # finalize
    wb.save(out)
    return out, lookup_info
