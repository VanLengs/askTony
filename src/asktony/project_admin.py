from __future__ import annotations

import csv
import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from rich.table import Table

from asktony.db import DB


@dataclass(frozen=True)
class ImportIssue:
    file: str
    row: int
    key: str
    field: str
    message: str


def issues_to_table(issues: list[ImportIssue], *, title: str = "Data Quality Issues") -> Table:
    table = Table(title=f"{title} (showing up to 50)")
    table.add_column("file")
    table.add_column("row", justify="right")
    table.add_column("key")
    table.add_column("field")
    table.add_column("message")
    for i in issues[:50]:
        table.add_row(i.file, str(i.row), i.key, i.field, i.message)
    return table


def _norm_id(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _norm_header(name: str) -> str:
    s = (name or "").strip()
    if not s:
        return ""
    if s.endswith("*"):
        s = s[:-1]
    s = s.strip().lower().replace(" ", "_")
    return s


def _pinyin_project_id(project_name: str) -> str:
    """
    Prefer OS-provided transliteration (macOS Foundation) to derive a stable pinyin id.
    Fallback to `pypinyin` if OS transliteration is unavailable.
    """
    try:
        from Foundation import NSMutableString  # type: ignore[import-not-found]
        from Foundation import CFStringTransform  # type: ignore[import-not-found]
    except Exception as e:  # noqa: BLE001
        try:
            from pypinyin import Style, lazy_pinyin  # type: ignore[import-not-found]

            parts = lazy_pinyin(project_name, style=Style.NORMAL)
            return _norm_id("".join(parts))
        except Exception as e2:  # noqa: BLE001
            raise RuntimeError(
                "OS transliteration unavailable; install pypinyin or provide explicit project_id"
            ) from e2

    s = NSMutableString.stringWithString_(project_name)
    # Mandarin Latin with tone marks, then strip diacritics.
    CFStringTransform(s, None, "Any-Latin", False)
    CFStringTransform(s, None, "Latin-ASCII", False)
    derived = _norm_id(str(s))
    if derived:
        return derived

    # Some environments may have Foundation but transliteration disabled; fallback.
    try:
        from pypinyin import Style, lazy_pinyin  # type: ignore[import-not-found]

        parts = lazy_pinyin(project_name, style=Style.NORMAL)
        return _norm_id("".join(parts))
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "OS transliteration produced empty id; install pypinyin or provide explicit project_id"
        ) from e


def _parse_date(s: str) -> dt.date | None:
    s = (s or "").strip()
    if not s:
        return None
    return dt.date.fromisoformat(s)


def _parse_float(s: str) -> float | None:
    s = (s or "").strip()
    if not s:
        return None
    return float(s)


def _ranges_overlap(a_start: dt.date, a_end: dt.date | None, b_start: dt.date, b_end: dt.date | None) -> bool:
    a2 = a_end or dt.date.max
    b2 = b_end or dt.date.max
    return not (a2 < b_start or b2 < a_start)


class ProjectAdmin:
    def __init__(self, db: DB) -> None:
        self.db = db

    def ensure_schema(self) -> None:
        with self.db.connect() as conn:
            conn.execute("CREATE SCHEMA IF NOT EXISTS gold")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS gold.dim_project (
                  project_id TEXT PRIMARY KEY,
                  project_name TEXT,
                  project_type TEXT,
                  status TEXT
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS gold.bridge_project_repo (
                  project_id TEXT,
                  repo_id TEXT,
                  start_at DATE,
                  end_at DATE,
                  weight DOUBLE,
                  PRIMARY KEY(project_id, repo_id, start_at)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS gold.bridge_project_person_role (
                  project_id TEXT,
                  employee_id TEXT,
                  project_role TEXT,
                  start_at DATE,
                  end_at DATE,
                  allocation DOUBLE,
                  PRIMARY KEY(project_id, employee_id, project_role, start_at)
                )
                """
            )

    def import_project_info(
        self,
        *,
        xlsx_file: Path | None = None,
        project_file: Path | None,
        project_repo_file: Path | None,
        project_member_file: Path | None,
        dry_run: bool = False,
    ) -> tuple[list[ImportIssue], dict[str, int]]:
        self.ensure_schema()
        issues: list[ImportIssue] = []
        stats: dict[str, int] = {
            "projects_upserted": 0,
            "project_repo_rows_upserted": 0,
            "project_member_rows_upserted": 0,
            "warnings_repo_weight_sum_outside_1": 0,
            "warnings_unknown_employee_id": 0,
        }

        projects: list[tuple[str, str | None, str | None, str | None]] = []
        repo_rows: list[tuple[str, str, dt.date, dt.date | None, float]] = []
        member_rows: list[tuple[str, str, str, dt.date, dt.date | None, float]] = []

        # Existing dim_project map (for mapping by project_name)
        existing_project_name_to_id: dict[str, str] = {}
        existing_project_id_set: set[str] = set()

        # load known employee_ids for warnings
        known_employee_ids: set[str] = set()
        full_name_to_employee_ids: dict[str, set[str]] = {}
        with self.db.connect() as conn:
            for pid, pname in conn.execute(
                "SELECT project_id, project_name FROM gold.dim_project"
            ).fetchall():
                if pid:
                    existing_project_id_set.add(_norm_id(str(pid)))
                if pname:
                    existing_project_name_to_id[str(pname).strip()] = _norm_id(str(pid))
            for (eid,) in conn.execute(
                """
                SELECT DISTINCT NULLIF(TRIM(employee_id), '') AS employee_id
                FROM gold.dim_member_enrichment
                WHERE employee_id IS NOT NULL AND employee_id <> ''
                """
            ).fetchall():
                if eid:
                    known_employee_ids.add(str(eid))
            for eid, full_name in conn.execute(
                """
                SELECT
                  NULLIF(TRIM(employee_id), '') AS employee_id,
                  NULLIF(TRIM(full_name), '') AS full_name
                FROM gold.dim_member_enrichment
                WHERE employee_id IS NOT NULL AND employee_id <> ''
                  AND full_name IS NOT NULL AND full_name <> ''
                """
            ).fetchall():
                if eid and full_name:
                    full_name_to_employee_ids.setdefault(str(full_name).strip(), set()).add(str(eid).strip())

        # In-import project name->id mapping (from sheet/CSV)
        project_name_to_id: dict[str, str] = {}

        def resolve_project_id(*, row_idx: int, file: str, project_id: str, project_name: str, key: str) -> str | None:
            pid = _norm_id(project_id)
            pname = (project_name or "").strip()
            if pid:
                return pid
            if pname:
                if pname in project_name_to_id:
                    return project_name_to_id[pname]
                if pname in existing_project_name_to_id:
                    return existing_project_name_to_id[pname]
                try:
                    derived = _pinyin_project_id(pname)
                    if not derived:
                        raise ValueError("invalid derived project_id")
                    return derived
                except Exception as e:  # noqa: BLE001
                    issues.append(ImportIssue(file, row_idx, key or pname, "project_id", str(e)))
                    return None
            issues.append(ImportIssue(file, row_idx, key, "project_id", "missing project_id/project_name"))
            return None

        def resolve_employee_id(*, row_idx: int, file: str, project_id: str, employee_id: str, full_name: str) -> str | None:
            eid = str(employee_id or "").strip()
            if eid:
                return eid
            name = str(full_name or "").strip()
            if not name:
                issues.append(ImportIssue(file, row_idx, project_id, "employee_id", "missing employee_id/full_name"))
                return None
            ids = sorted(full_name_to_employee_ids.get(name, set()))
            if not ids:
                issues.append(ImportIssue(file, row_idx, f"{project_id}:{name}", "employee_id", "full_name not found"))
                return None
            if len(ids) > 1:
                issues.append(
                    ImportIssue(
                        file,
                        row_idx,
                        f"{project_id}:{name}",
                        "employee_id",
                        f"full_name is ambiguous (employee_id candidates: {ids})",
                    )
                )
                return None
            return ids[0]

        if xlsx_file is not None:
            xlsx = xlsx_file.expanduser()
            try:
                from openpyxl import load_workbook  # type: ignore[import-not-found]
            except Exception as e:  # noqa: BLE001
                raise ModuleNotFoundError(
                    "Missing dependency 'openpyxl'. Install it (e.g. `pip install openpyxl`) to import XLSX."
                ) from e

            wb = load_workbook(xlsx, data_only=True)
            name_map = {str(n).strip().lower(): n for n in wb.sheetnames}

            def read_sheet(sheet_key: str) -> list[dict[str, Any]]:
                actual = name_map.get(sheet_key.lower())
                if actual is None:
                    return []
                ws = wb[actual]
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    return []
                headers = [_norm_header("" if h is None else str(h)) for h in header_row]
                out_rows: list[dict[str, Any]] = []
                for r in rows_iter:
                    if r is None:
                        continue
                    if all(v is None or str(v).strip() == "" for v in r):
                        continue
                    d: dict[str, Any] = {}
                    for k, v in zip(headers, r, strict=False):
                        if not k:
                            continue
                        if isinstance(v, (dt.datetime, dt.date)):
                            d[k] = v.date().isoformat() if isinstance(v, dt.datetime) else v.isoformat()
                        else:
                            d[k] = "" if v is None else str(v).strip()
                    out_rows.append(d)
                return out_rows

            # Convert xlsx sheets into the same dict-row structure as CSV readers.
            project_rows = read_sheet("dim_project")
            project_repo_rows = read_sheet("bridge_project_repo")
            project_member_rows = read_sheet("bridge_project_person_role")

            # Feed into the normal loaders by writing into in-memory-like loops.
            if project_rows:
                for idx, row in enumerate(project_rows, start=2):
                    project_name = str(row.get("project_name") or "").strip()
                    if not project_name:
                        issues.append(ImportIssue("project", idx, "", "project_name", "missing project_name"))
                        continue
                    project_id = str(row.get("project_id") or "").strip()
                    if not project_id:
                        try:
                            project_id = _pinyin_project_id(project_name)
                        except Exception as e:  # noqa: BLE001
                            issues.append(ImportIssue("project", idx, project_name, "project_id", str(e)))
                            continue
                    project_id = _norm_id(project_id)
                    if not project_id:
                        issues.append(ImportIssue("project", idx, project_name, "project_id", "invalid project_id"))
                        continue
                    project_name_to_id[project_name] = project_id
                    project_type = str(row.get("project_type") or "").strip() or None
                    status = str(row.get("status") or "").strip() or None
                    projects.append((project_id, project_name, project_type, status))

            if project_repo_rows:
                for idx, row in enumerate(project_repo_rows, start=2):
                    project_id = str(row.get("project_id") or "")
                    project_name = str(row.get("project_name") or "").strip()
                    repo_id = str(row.get("repo_id") or "").strip()
                    resolved_pid = resolve_project_id(
                        row_idx=idx,
                        file="project_repo",
                        project_id=project_id,
                        project_name=project_name,
                        key=repo_id,
                    )
                    if not resolved_pid:
                        continue
                    if not repo_id:
                        issues.append(ImportIssue("project_repo", idx, resolved_pid, "repo_id", "missing repo_id"))
                        continue
                    try:
                        start_at = _parse_date(str(row.get("start_at") or ""))
                        end_at = _parse_date(str(row.get("end_at") or ""))
                        if start_at is None:
                            raise ValueError("missing start_at")
                        if end_at is not None and end_at < start_at:
                            raise ValueError("end_at < start_at")
                    except Exception as e:  # noqa: BLE001
                        issues.append(ImportIssue("project_repo", idx, f"{resolved_pid}:{repo_id}", "date", str(e)))
                        continue
                    try:
                        weight = _parse_float(str(row.get("weight") or "")) or 1.0
                    except Exception as e:  # noqa: BLE001
                        issues.append(
                            ImportIssue("project_repo", idx, f"{resolved_pid}:{repo_id}", "weight", str(e))
                        )
                        continue
                    if not (0.0 < weight <= 1.0):
                        issues.append(
                            ImportIssue(
                                "project_repo",
                                idx,
                                f"{resolved_pid}:{repo_id}",
                                "weight",
                                "weight must be in (0, 1]",
                            )
                        )
                        continue
                    repo_rows.append((resolved_pid, repo_id, start_at, end_at, float(weight)))

            if project_member_rows:
                for idx, row in enumerate(project_member_rows, start=2):
                    project_id = str(row.get("project_id") or "")
                    project_name = str(row.get("project_name") or "").strip()
                    employee_id = str(row.get("employee_id") or "").strip()
                    full_name = str(row.get("full_name") or "").strip()
                    resolved_pid = resolve_project_id(
                        row_idx=idx,
                        file="project_member",
                        project_id=project_id,
                        project_name=project_name,
                        key=employee_id,
                    )
                    if not resolved_pid:
                        continue
                    resolved_eid = resolve_employee_id(
                        row_idx=idx, file="project_member", project_id=resolved_pid, employee_id=employee_id, full_name=full_name
                    )
                    if not resolved_eid:
                        continue
                    project_role = str(row.get("project_role") or "member").strip() or "member"
                    try:
                        start_at = _parse_date(str(row.get("start_at") or ""))
                        end_at = _parse_date(str(row.get("end_at") or ""))
                        if start_at is None:
                            raise ValueError("missing start_at")
                        if end_at is not None and end_at < start_at:
                            raise ValueError("end_at < start_at")
                    except Exception as e:  # noqa: BLE001
                        issues.append(
                            ImportIssue("project_member", idx, f"{resolved_pid}:{employee_id}", "date", str(e))
                        )
                        continue
                    try:
                        allocation = _parse_float(str(row.get("allocation") or "")) or 1.0
                    except Exception as e:  # noqa: BLE001
                        issues.append(
                            ImportIssue(
                                "project_member",
                                idx,
                                f"{resolved_pid}:{employee_id}",
                                "allocation",
                                str(e),
                            )
                        )
                        continue
                    if not (0.0 < float(allocation) <= 1.0):
                        issues.append(
                            ImportIssue(
                                "project_member",
                                idx,
                                f"{resolved_pid}:{employee_id}",
                                "allocation",
                                "allocation must be in (0, 1]",
                            )
                        )
                        continue
                    if known_employee_ids and employee_id not in known_employee_ids:
                        stats["warnings_unknown_employee_id"] += 1
                    member_rows.append(
                        (resolved_pid, resolved_eid, project_role, start_at, end_at, float(allocation))
                    )

            # XLSX mode ignores CSV inputs
            project_file = None
            project_repo_file = None
            project_member_file = None

        if project_file is not None:
            p = project_file.expanduser()
            with p.open("r", encoding="utf-8-sig", newline="") as f:
                r = csv.DictReader(f)
                required = {"project_name"}
                if not required.issubset(set(r.fieldnames or [])):
                    raise ValueError(f"project_file missing required columns: {sorted(required)}")
                for idx, row in enumerate(r, start=2):
                    project_name = str(row.get("project_name") or "").strip()
                    if not project_name:
                        issues.append(ImportIssue("project", idx, "", "project_name", "missing project_name"))
                        continue
                    project_id = str(row.get("project_id") or "").strip()
                    if not project_id:
                        try:
                            project_id = _pinyin_project_id(project_name)
                        except Exception as e:  # noqa: BLE001
                            issues.append(
                                ImportIssue("project", idx, project_name, "project_id", str(e))
                            )
                            continue
                    project_id = _norm_id(project_id)
                    if not project_id:
                        issues.append(
                            ImportIssue("project", idx, project_name, "project_id", "invalid project_id")
                        )
                        continue
                    project_name_to_id[project_name] = project_id
                    project_type = str(row.get("project_type") or "").strip() or None
                    status = str(row.get("status") or "").strip() or None
                    projects.append((project_id, project_name, project_type, status))

        if project_repo_file is not None:
            p = project_repo_file.expanduser()
            with p.open("r", encoding="utf-8-sig", newline="") as f:
                r = csv.DictReader(f)
                required = {"repo_id", "start_at"}
                if not required.issubset(set(r.fieldnames or [])):
                    raise ValueError(f"project_repo_file missing required columns: {sorted(required)}")
                for idx, row in enumerate(r, start=2):
                    project_id = str(row.get("project_id") or "")
                    project_name = str(row.get("project_name") or "").strip()
                    repo_id = str(row.get("repo_id") or "").strip()
                    resolved_pid = resolve_project_id(
                        row_idx=idx,
                        file="project_repo",
                        project_id=project_id,
                        project_name=project_name,
                        key=repo_id,
                    )
                    if not resolved_pid:
                        continue
                    if not repo_id:
                        issues.append(ImportIssue("project_repo", idx, resolved_pid, "repo_id", "missing repo_id"))
                        continue
                    try:
                        start_at = _parse_date(str(row.get("start_at") or ""))
                        end_at = _parse_date(str(row.get("end_at") or ""))
                        if start_at is None:
                            raise ValueError("missing start_at")
                        if end_at is not None and end_at < start_at:
                            raise ValueError("end_at < start_at")
                    except Exception as e:  # noqa: BLE001
                        issues.append(ImportIssue("project_repo", idx, f"{resolved_pid}:{repo_id}", "date", str(e)))
                        continue
                    try:
                        weight = _parse_float(str(row.get("weight") or "")) or 1.0
                    except Exception as e:  # noqa: BLE001
                        issues.append(
                            ImportIssue("project_repo", idx, f"{resolved_pid}:{repo_id}", "weight", str(e))
                        )
                        continue
                    if not (0.0 < weight <= 1.0):
                        issues.append(
                            ImportIssue(
                                "project_repo",
                                idx,
                                f"{resolved_pid}:{repo_id}",
                                "weight",
                                "weight must be in (0, 1]",
                            )
                        )
                        continue
                    repo_rows.append((resolved_pid, repo_id, start_at, end_at, float(weight)))

        if project_member_file is not None:
            p = project_member_file.expanduser()
            with p.open("r", encoding="utf-8-sig", newline="") as f:
                r = csv.DictReader(f)
                required = {"employee_id", "start_at"}
                if not required.issubset(set(r.fieldnames or [])):
                    raise ValueError(f"project_member_file missing required columns: {sorted(required)}")
                for idx, row in enumerate(r, start=2):
                    project_id = str(row.get("project_id") or "")
                    project_name = str(row.get("project_name") or "").strip()
                    employee_id = str(row.get("employee_id") or "").strip()
                    full_name = str(row.get("full_name") or "").strip()
                    resolved_pid = resolve_project_id(
                        row_idx=idx,
                        file="project_member",
                        project_id=project_id,
                        project_name=project_name,
                        key=employee_id,
                    )
                    if not resolved_pid:
                        continue
                    resolved_eid = resolve_employee_id(
                        row_idx=idx,
                        file="project_member",
                        project_id=resolved_pid,
                        employee_id=employee_id,
                        full_name=full_name,
                    )
                    if not resolved_eid:
                        continue
                    project_role = str(row.get("project_role") or "member").strip() or "member"
                    project_role = project_role.strip()
                    try:
                        start_at = _parse_date(str(row.get("start_at") or ""))
                        end_at = _parse_date(str(row.get("end_at") or ""))
                        if start_at is None:
                            raise ValueError("missing start_at")
                        if end_at is not None and end_at < start_at:
                            raise ValueError("end_at < start_at")
                    except Exception as e:  # noqa: BLE001
                        issues.append(
                            ImportIssue("project_member", idx, f"{resolved_pid}:{employee_id}", "date", str(e))
                        )
                        continue
                    try:
                        allocation = _parse_float(str(row.get("allocation") or "")) or 1.0
                    except Exception as e:  # noqa: BLE001
                        issues.append(
                            ImportIssue(
                                "project_member",
                                idx,
                                f"{resolved_pid}:{employee_id}",
                                "allocation",
                                str(e),
                            )
                        )
                        continue
                    if not (0.0 < float(allocation) <= 1.0):
                        issues.append(
                            ImportIssue(
                                "project_member",
                                idx,
                                f"{resolved_pid}:{employee_id}",
                                "allocation",
                                "allocation must be in (0, 1]",
                            )
                        )
                        continue
                    if known_employee_ids and employee_id not in known_employee_ids:
                        stats["warnings_unknown_employee_id"] += 1
                    member_rows.append(
                        (resolved_pid, resolved_eid, project_role, start_at, end_at, float(allocation))
                    )

        # Overlap checks (in-import)
        def check_overlaps_repo() -> None:
            by_key: dict[tuple[str, str], list[tuple[dt.date, dt.date | None, float]]] = {}
            for pid, rid, s, e, w in repo_rows:
                by_key.setdefault((pid, rid), []).append((s, e, w))
            for (pid, rid), ranges in by_key.items():
                ranges_sorted = sorted(ranges, key=lambda x: x[0])
                for i in range(1, len(ranges_sorted)):
                    a_s, a_e, _w1 = ranges_sorted[i - 1]
                    b_s, b_e, _w2 = ranges_sorted[i]
                    if _ranges_overlap(a_s, a_e, b_s, b_e):
                        issues.append(
                            ImportIssue(
                                "project_repo",
                                0,
                                f"{pid}:{rid}",
                                "date",
                                "overlapping ranges within (project_id, repo_id)",
                            )
                        )
                        return

        def check_overlaps_member() -> None:
            by_key: dict[tuple[str, str, str], list[tuple[dt.date, dt.date | None, float]]] = {}
            for pid, eid, role, s, e, a in member_rows:
                by_key.setdefault((pid, eid, role.strip().lower()), []).append((s, e, a))
            for (pid, eid, role), ranges in by_key.items():
                ranges_sorted = sorted(ranges, key=lambda x: x[0])
                for i in range(1, len(ranges_sorted)):
                    a_s, a_e, _a1 = ranges_sorted[i - 1]
                    b_s, b_e, _a2 = ranges_sorted[i]
                    if _ranges_overlap(a_s, a_e, b_s, b_e):
                        issues.append(
                            ImportIssue(
                                "project_member",
                                0,
                                f"{pid}:{eid}:{role}",
                                "date",
                                "overlapping ranges within (project_id, employee_id, project_role)",
                            )
                        )
                        return

        check_overlaps_repo()
        check_overlaps_member()

        # Repo weight sum warnings (best-effort): if multiple project mappings overlap for same repo,
        # flag cases where sum(weight) is far from 1 for the overlapping period.
        by_repo: dict[str, list[tuple[str, dt.date, dt.date | None, float]]] = {}
        for pid, rid, s, e, w in repo_rows:
            by_repo.setdefault(rid, []).append((pid, s, e, w))
        for rid, ranges in by_repo.items():
            if len(ranges) <= 1:
                continue
            # check pairwise overlaps; if overlaps exist, compute sum of weights among overlapping rows
            for i in range(len(ranges)):
                for j in range(i + 1, len(ranges)):
                    _pid1, s1, e1, w1 = ranges[i]
                    _pid2, s2, e2, w2 = ranges[j]
                    if _ranges_overlap(s1, e1, s2, e2):
                        total = w1 + w2
                        if not (0.95 <= total <= 1.05):
                            stats["warnings_repo_weight_sum_outside_1"] += 1
                            break

        if issues:
            return issues, stats

        if dry_run:
            return issues, stats

        with self.db.connect() as conn:
            if projects:
                conn.executemany(
                    """
                    INSERT INTO gold.dim_project(project_id, project_name, project_type, status)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT (project_id) DO UPDATE SET
                      project_name = EXCLUDED.project_name,
                      project_type = EXCLUDED.project_type,
                      status = EXCLUDED.status
                    """,
                    projects,
                )
                stats["projects_upserted"] = len(projects)

            if repo_rows:
                conn.executemany(
                    """
                    INSERT INTO gold.bridge_project_repo(project_id, repo_id, start_at, end_at, weight)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT (project_id, repo_id, start_at) DO UPDATE SET
                      end_at = EXCLUDED.end_at,
                      weight = EXCLUDED.weight
                    """,
                    repo_rows,
                )
                stats["project_repo_rows_upserted"] = len(repo_rows)

            if member_rows:
                conn.executemany(
                    """
                    INSERT INTO gold.bridge_project_person_role(
                      project_id, employee_id, project_role, start_at, end_at, allocation
                    )
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT (project_id, employee_id, project_role, start_at) DO UPDATE SET
                      end_at = EXCLUDED.end_at,
                      allocation = EXCLUDED.allocation
                    """,
                    member_rows,
                )
                stats["project_member_rows_upserted"] = len(member_rows)

        return issues, stats
