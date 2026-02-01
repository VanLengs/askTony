from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

import pandas as pd
import typer
from rich.console import Console

from asktony.config import load_config
from asktony.warehouse import Warehouse

universe_app = typer.Typer(help="一键导出 Universe（XLSX）。", add_completion=False)
console = Console()


def _xlsx_col_letter(n: int) -> str:
    # 1-based
    out = ""
    while n:
        n, r = divmod(n - 1, 26)
        out = chr(65 + r) + out
    return out


def _write_table(ws, columns: list[str], rows: list[tuple], *, hide_columns: set[str]) -> None:
    from openpyxl.styles import Font

    header_font = Font(bold=True)
    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font

    for r in rows:
        masked_row = []
        for col, v in zip(columns, r, strict=False):
            if col in hide_columns:
                masked_row.append("***")
            else:
                masked_row.append("" if v is None else v)
        # If row is shorter than columns (shouldn't happen), pad.
        if len(masked_row) < len(columns):
            masked_row.extend([""] * (len(columns) - len(masked_row)))
        ws.append(masked_row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_xlsx_col_letter(len(columns))}{max(1, len(rows) + 1)}"

    # widths + hide
    for i, name in enumerate(columns, start=1):
        col_letter = _xlsx_col_letter(i)
        ws.column_dimensions[col_letter].width = min(max(len(str(name)) + 2, 12), 40)
        if name in hide_columns:
            ws.column_dimensions[col_letter].hidden = True


def _add_image_sheet(wb, *, title: str, image_path: Path) -> None:
    ws = wb.create_sheet(title=title)
    ws["A1"] = str(image_path)
    try:
        from openpyxl.drawing.image import Image
    except Exception:  # noqa: BLE001
        return

    try:
        img = Image(str(image_path))
        img.anchor = "A3"
        ws.add_image(img)
    except Exception:  # noqa: BLE001
        # Leave path only if embedding fails.
        return


@universe_app.command("export")
def export(
    months: int = typer.Option(2, min=1, max=60, help="最近 N 个月（分析窗口）"),
    output: Path = typer.Option(Path("output/universe.xlsx"), help="输出 XLSX 文件路径"),
    hide_sensitive: bool = typer.Option(
        False,
        "--hide-sensitive",
        help="隐藏敏感字段列（age/job_rank/education_level/collodge/major）",
    ),
    include_images: bool = typer.Option(True, "--include-images/--no-images", help="是否生成并嵌入图片"),
) -> None:
    """
    导出一个 XLSX：每个分析表/图片单独一个 sheet。

    - 表：active-repos, member-commits, repo-member-commits, active-members, inactive-members, line-manager-dev-activity
    - 表（新增）：active-employee-score, suspicious-committers
    - 图：report.png, line_manager_dev_activity_radar.png, active_employee_score_radar.png, anti_fraud_report.png
    """
    try:
        import openpyxl  # noqa: F401
    except Exception as e:  # noqa: BLE001
        raise typer.BadParameter(
            "missing dependency openpyxl; please install project dependencies first"
        ) from e

    from openpyxl import Workbook

    cfg = load_config()
    wh = Warehouse.from_config(cfg)

    hide_columns = set()
    if hide_sensitive:
        hide_columns = {"age", "job_rank", "education_level", "collodge", "major"}

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    def add_table_sheet(name: str, columns: list[str], rows: list[tuple]) -> None:
        ws = wb.create_sheet(title=name)
        _write_table(ws, columns, rows, hide_columns=hide_columns)

    columns, rows = wh.active_repos_data(months=months, top=None)
    add_table_sheet("active-repos", columns, rows)

    columns, rows = wh.member_commits_all_repos_data(months=months, top=None)
    add_table_sheet("member-commits", columns, rows)

    columns, rows = wh.repo_member_commits_data(months=months, top=None)
    add_table_sheet("repo-member-commits", columns, rows)

    columns, rows = wh.active_members_data(months=months, top=None, all_fields=True)
    add_table_sheet("active-members", columns, rows)

    columns, rows = wh.inactive_members_data(months=months, top=None, all_fields=True)
    add_table_sheet("inactive-members", columns, rows)

    lm_columns, lm_rows = wh.line_manager_dev_activity_data(months=months, top=None)
    add_table_sheet("line-manager-dev-activity", lm_columns, lm_rows)

    emp_columns, emp_rows = wh.active_employee_score_data(months=months, top=None)
    add_table_sheet("active-employee-score", emp_columns, emp_rows)

    af_columns, af_rows = wh.suspicious_committers_data(months=months, top=None)
    add_table_sheet("suspicious-committers", af_columns, af_rows)

    report_png = Path("output/report.png")
    radar_png = Path("output/line_manager_dev_activity_radar.png")
    emp_radar_png = Path("output/active_employee_score_radar.png")
    anti_fraud_png = Path("output/anti_fraud_report.png")

    if include_images:
        since_dt = dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=30 * months)
        try:
            from asktony.visualize import (
                plot_active_employee_score_radar,
                plot_anti_fraud_report,
                plot_line_manager_dev_activity_radar,
                plot_report,
            )

            # report.png
            report_png.parent.mkdir(parents=True, exist_ok=True)
            plot_report(
                db_path=cfg.db_path_resolved,
                since_dt=since_dt,
                top_n=10,
                bottom_n=10,
                output=report_png,
                dpi=180,
            )

            # radar.png
            radar_png.parent.mkdir(parents=True, exist_ok=True)
            df = pd.DataFrame(lm_rows, columns=lm_columns)
            plot_line_manager_dev_activity_radar(df=df, output=radar_png, dpi=200, top_n=12, bottom_n=6)

            # active employee radar.png
            emp_radar_png.parent.mkdir(parents=True, exist_ok=True)
            emp_df = pd.DataFrame(emp_rows, columns=emp_columns)
            plot_active_employee_score_radar(df=emp_df, output=emp_radar_png, dpi=200, top_n=10, bottom_n=10)

            # anti-fraud report.png
            anti_fraud_png.parent.mkdir(parents=True, exist_ok=True)
            af_df = pd.DataFrame(af_rows, columns=af_columns)
            plot_anti_fraud_report(df=af_df, output=anti_fraud_png, dpi=200, top_n=10)
        except Exception:  # noqa: BLE001
            # If plotting fails (deps/fonts), still export tables and add placeholders.
            pass

    if report_png.exists():
        _add_image_sheet(wb, title="report.png", image_path=report_png)
    else:
        ws = wb.create_sheet(title="report.png")
        ws["A1"] = "report.png not found (run: asktony visualize report)"

    # Sheet title max length is 31.
    radar_sheet = "line_mgr_radar.png"
    if radar_png.exists():
        _add_image_sheet(wb, title=radar_sheet, image_path=radar_png)
    else:
        ws = wb.create_sheet(title=radar_sheet)
        ws["A1"] = "line_manager_dev_activity_radar.png not found (run: asktony visualize line-manager-dev-activity)"

    emp_radar_sheet = "emp_score_radar.png"
    if emp_radar_png.exists():
        _add_image_sheet(wb, title=emp_radar_sheet, image_path=emp_radar_png)
    else:
        ws = wb.create_sheet(title=emp_radar_sheet)
        ws["A1"] = "active_employee_score_radar.png not found (run: asktony visualize active-employee-score)"

    anti_fraud_sheet = "anti_fraud.png"
    if anti_fraud_png.exists():
        _add_image_sheet(wb, title=anti_fraud_sheet, image_path=anti_fraud_png)
    else:
        ws = wb.create_sheet(title=anti_fraud_sheet)
        ws["A1"] = "anti_fraud_report.png not found (run: asktony visualize anti-fraud-report)"

    out = output.expanduser()
    today = dt.date.today().isoformat()  # YYYY-MM-DD
    # Always append date suffix unless user already did.
    if out.suffix.lower() == ".xlsx":
        if not re.search(r"(\d{4}-\d{2}-\d{2}|\d{8})$", out.stem):
            out = out.with_name(f"{out.stem}_{today}{out.suffix}")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    console.print({"output": str(out), "date_suffix": today})
